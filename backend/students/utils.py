import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


def export_students_excel(queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1E40AF")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill("solid", fgColor="EFF6FF")

    headers = ["S.No", "Name", "Father's Name", "Phone Number", "Class", "Village", "Registered On"]
    col_widths = [6, 24, 24, 16, 18, 20, 22]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    for row_idx, student in enumerate(queryset, start=2):
        row_data = [
            row_idx - 1,
            student.name,
            student.father_name,
            student.phone_number,
            student.get_class_enrolled_display(),
            student.village.name if student.village else '—',
            student.created_at.strftime("%d %b %Y") if student.created_at else '',
        ]
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="students_data.xlsx"'
    return response
